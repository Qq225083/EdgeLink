import ipaddress
import json
from datetime import datetime
from typing import Any, Union

from sqlalchemy import func, select as sa_select, update as sa_update
from sqlalchemy.ext.asyncio import AsyncSession

from common.vo import CrudResponseModel, PageModel
from exceptions.exception import ServiceException
from module_plc.dao.device_dao import DeviceDao
from module_plc.dao.monitor_dao import MonitorDao
from module_plc.dao.tag_dao import TagDao
from module_plc.entity.do.change_log_do import PlcChangeLog
from module_plc.entity.do.device_do import PlcDevice
from module_plc.entity.do.monitor_do import DeviceCommStatus
from module_plc.entity.do.tag_do import PlcTag
from module_plc.entity.vo.device_vo import DeletePlcDeviceModel, PlcDeviceModel, PlcDevicePageQueryModel
from module_plc.entity.vo.tag_vo import TagModel, TagPageQueryModel
from module_plc.service.driver_service import DriverService
from utils.common_util import CamelCaseUtil
from utils.log_util import logger


class DeviceService:
    """
    PLC设备模块服务层
    """

    # FX系列不允许使用的帧格式
    FX_FORBIDDEN_FRAMES = {'4E'}

    @staticmethod
    def _json_safe(value: dict | None) -> dict | None:
        """将 dict 转为 JSON 可序列化形式（datetime 等类型转字符串），避免 JSON 列写入失败"""
        if value is None:
            return None
        try:
            return json.loads(json.dumps(value, ensure_ascii=False, default=str))
        except (TypeError, ValueError):
            return {'_raw': str(value)}

    @classmethod
    async def _log_change(
        cls, db: AsyncSession, change_type: str, target_type: str,
        target_id: int, target_name: str = '', change_by: str = '',
        before_value: dict | None = None, after_value: dict | None = None,
    ) -> None:
        """记录修改履历（字段截断到列长，防止 commit 时 flush 失败毒化主事务）"""
        try:
            log = PlcChangeLog(
                change_type=change_type[:20],
                target_type=target_type[:20],
                target_id=target_id,
                target_name=(target_name or '')[:100],
                change_by=(change_by or '')[:64],
                before_value=cls._json_safe(before_value),
                after_value=cls._json_safe(after_value),
            )
            db.add(log)
        except Exception as exc:
            logger.warning(f'修改履历记录失败: {exc}')

    @classmethod
    def _validate_name_code_length(cls, device_name: str, device_code: str) -> None:
        """
        校验设备名称/编号长度（对应 plc_device 列宽 device_name varchar(100) / device_code varchar(50)），
        防止超长值直达数据库抛 DataError(1406) —— 克隆后缀叠加是常见来源

        :param device_name: 设备名称
        :param device_code: 设备编号
        :raises ServiceException: 超长时抛出
        """
        if device_name and len(device_name) > 100:
            raise ServiceException(message=f'设备名称过长（{len(device_name)}字符），最大允许100字符')
        if device_code and len(device_code) > 50:
            raise ServiceException(message=f'设备编号过长（{len(device_code)}字符），最大允许50字符')

    @classmethod
    def _validate_ip(cls, ip_str: str, allow_port: bool = False) -> None:
        """
        校验IP地址格式是否合法

        :param ip_str: IP地址字符串，支持纯 IP 或 IP:端口（当 allow_port=True 时）
        :param allow_port: 是否允许 IP:端口 格式（用于采集节点标识）
        :raises ServiceException: IP格式不正确时抛出
        """
        if not ip_str:
            return
        # 支持 IP:端口 格式（多实例采集节点标识）
        if allow_port and ':' in ip_str:
            parts = ip_str.rsplit(':', 1)
            if len(parts) == 2:
                ip_part, port_part = parts
                try:
                    ipaddress.ip_address(ip_part)
                except ValueError:
                    raise ServiceException(message=f'IP地址格式不正确：{ip_str}')
                if not port_part.isdigit() or not (1 <= int(port_part) <= 65535):
                    raise ServiceException(message=f'端口范围不正确：{port_part}，有效范围为1-65535')
                return
        try:
            ipaddress.ip_address(ip_str)
        except ValueError:
            raise ServiceException(message=f'IP地址格式不正确：{ip_str}')

    @classmethod
    def _validate_port(cls, port: int) -> None:
        """
        校验端口号范围是否合法（1-65535）

        :param port: 端口号
        :raises ServiceException: 端口范围不正确时抛出
        """
        if port < 1 or port > 65535:
            raise ServiceException(message=f'端口号范围不正确：{port}，有效范围为1-65535')

    @classmethod
    def _validate_series_frame(cls, plc_series: str, mc_frame: str) -> None:
        """
        校验PLC系列与MC帧格式的兼容性
        FX系列不允许使用4E帧

        :param plc_series: PLC系列
        :param mc_frame: MC帧格式
        :raises ServiceException: 不兼容时抛出
        """
        if plc_series and plc_series.upper() == 'FX' and mc_frame and mc_frame in cls.FX_FORBIDDEN_FRAMES:
            raise ServiceException(
                message=f'FX系列PLC不支持{mc_frame}帧格式，请选择3E帧或更换PLC系列'
            )

    @classmethod
    def _require_plc_ip_for_com_type(cls, plc_ip: str | None, com_type: str | None) -> None:
        """
        RS-232C 串口通信不需要 IP，其余方式必须有 PLC IP

        :param plc_ip: PLC IP地址
        :param com_type: 通信方式
        :raises ServiceException: 非RS-232C但缺少PLC IP时抛出
        """
        if com_type and com_type != 'PLC_RS232C' and not plc_ip:
            raise ServiceException(message=f'通信方式为 {com_type} 时必须填写 PLC IP 地址')

    @classmethod
    def _require_host_pc_ip(cls, host_pc_ip: str | None) -> None:
        """
        办公网IP 为必填项，用于确定由哪台采集PC负责此设备

        :param host_pc_ip: 办公网IP地址
        :raises ServiceException: host_pc_ip 为空时抛出
        """
        if not host_pc_ip:
            raise ServiceException(message='办公网IP为必填项，用于确定由哪台采集电脑负责此设备')

    @classmethod
    def _require_indust_net_ip(cls, mes_ip: str | None) -> None:
        """
        工业内网IP 为必填项，PLC所在的工业网络IP地址

        :param mes_ip: 工业内网IP地址
        :raises ServiceException: mes_ip 为空时抛出
        """
        if not mes_ip:
            raise ServiceException(message='工业内网IP为必填项，用于标识PLC所在的工业网络地址')

    # ==================== 设备 CRUD ====================

    @classmethod
    async def get_device_list_services(
        cls, query_db: AsyncSession, query_object: PlcDevicePageQueryModel, is_page: bool = False
    ) -> Union[PageModel, list[dict[str, Any]]]:
        """
        获取未删除的PLC设备列表信息（含点位数量）

        :param query_db: orm对象
        :param query_object: 查询参数对象
        :param is_page: 是否开启分页
        :return: PLC设备列表信息对象
        """
        device_list_result = await DeviceDao.get_device_list(query_db, query_object, is_page)

        if is_page and hasattr(device_list_result, 'rows') and device_list_result.rows:
            # 批量查询每个设备的点位数量
            device_ids = [row.get('id') for row in device_list_result.rows if isinstance(row, dict) and row.get('id')]
            if device_ids:
                counts_result = await query_db.execute(
                    sa_select(PlcTag.device_id, func.count(PlcTag.id))
                    .where(PlcTag.device_id.in_(device_ids), PlcTag.del_flag == '0')
                    .group_by(PlcTag.device_id)
                )
                counts = {row[0]: row[1] for row in counts_result}
                for row in device_list_result.rows:
                    if isinstance(row, dict):
                        row['tagCount'] = counts.get(row.get('id'), 0)

        return device_list_result

    @classmethod
    async def add_device_services(
        cls, query_db: AsyncSession, page_object: PlcDeviceModel
    ) -> CrudResponseModel:
        """
        新增PLC设备

        :param query_db: orm对象
        :param page_object: 新增PLC设备对象
        :return: 新增结果
        """
        # 校验 PLC IP：非RS-232C通信方式必须填写
        cls._require_plc_ip_for_com_type(page_object.plc_ip, page_object.com_type)
        # 校验 办公网IP：必填
        cls._require_host_pc_ip(page_object.host_pc_ip)
        if page_object.plc_ip:
            cls._validate_ip(page_object.plc_ip)
        cls._validate_ip(page_object.host_pc_ip, allow_port=True)
        # 校验 工业内网IP：必填
        cls._require_indust_net_ip(page_object.mes_ip)
        if page_object.mes_ip:
            cls._validate_ip(page_object.mes_ip)
        if page_object.backup_pc_ip:
            cls._validate_ip(page_object.backup_pc_ip, allow_port=True)
        if page_object.plc_port is not None:
            cls._validate_port(page_object.plc_port)
        if page_object.mes_port and page_object.mes_port > 0:
            cls._validate_port(page_object.mes_port)
        if page_object.plc_series and page_object.mc_frame:
            cls._validate_series_frame(page_object.plc_series, page_object.mc_frame)

        # 自动填充 driver_code（若前端未传，根据品牌/系列/通信方式映射）
        await DriverService.fill_driver_code_for_device_vo(query_db, page_object)
        # 校验 driver_code 有效且驱动已启用
        await DriverService.validate_driver_code(query_db, page_object.driver_code)

        # 名称/编号长度校验（先于唯一性查询，避免超长值直达数据库）
        cls._validate_name_code_length(page_object.device_name, page_object.device_code)

        # 设备名称唯一性校验
        if page_object.device_name:
            existing_by_name = await DeviceDao.get_device_by_name(query_db, page_object.device_name)
            if existing_by_name:
                raise ServiceException(message=f'设备名称已存在：{page_object.device_name}')

        # 设备编号唯一性校验
        if page_object.device_code:
            existing_by_code = await DeviceDao.get_device_by_code(query_db, page_object.device_code)
            if existing_by_code:
                raise ServiceException(message=f'设备编号已存在：{page_object.device_code}')

        try:
            db_device = await DeviceDao.add_device_dao(query_db, page_object)
            device_name = db_device.device_name
            await cls._log_change(
                query_db, 'add', 'device', db_device.id, device_name,
                change_by=page_object.update_by or '',
                after_value={'deviceName': device_name, 'deviceCode': db_device.device_code}
            )
            await query_db.commit()
            logger.info(f'新增PLC设备成功：{device_name}')
            return CrudResponseModel(is_success=True, message='新增成功（请手动发布配置后采集节点方可生效）')
        except Exception as e:
            await query_db.rollback()
            logger.exception(f'新增PLC设备失败：{e}')
            raise

    @classmethod
    async def edit_device_services(
        cls, query_db: AsyncSession, page_object: PlcDeviceModel
    ) -> CrudResponseModel:
        """
        编辑PLC设备信息

        :param query_db: orm对象
        :param page_object: 编辑PLC设备对象
        :return: 编辑结果
        """
        device_info = await DeviceDao.get_device_detail_by_id(query_db, page_object.id)
        if not device_info:
            raise ServiceException(message='PLC设备不存在')

        if page_object.plc_ip:
            cls._validate_ip(page_object.plc_ip)
        # 校验「最终态」：合并新旧值后，非 RS-232C 通信方式必须持有有效 PLC IP，办公网IP必填
        final_plc_ip = page_object.plc_ip if page_object.plc_ip is not None else device_info.plc_ip
        final_com_type = page_object.com_type if page_object.com_type else device_info.com_type
        final_host_pc_ip = page_object.host_pc_ip if page_object.host_pc_ip else device_info.host_pc_ip
        final_mes_ip = page_object.mes_ip if page_object.mes_ip else device_info.mes_ip
        cls._require_plc_ip_for_com_type(final_plc_ip, final_com_type)
        cls._require_host_pc_ip(final_host_pc_ip)
        cls._require_indust_net_ip(final_mes_ip)
        if page_object.host_pc_ip:
            cls._validate_ip(page_object.host_pc_ip, allow_port=True)
        if page_object.backup_pc_ip:
            cls._validate_ip(page_object.backup_pc_ip, allow_port=True)
        if page_object.mes_ip:
            cls._validate_ip(page_object.mes_ip)
        if page_object.plc_port is not None:
            cls._validate_port(page_object.plc_port)
        if page_object.mes_port and page_object.mes_port > 0:
            cls._validate_port(page_object.mes_port)
        # 校验最终态的 series + frame 兼容性（防止仅修改 series 为 FX 但现有 frame 为 4E 等场景）
        final_series = page_object.plc_series if page_object.plc_series else device_info.plc_series
        final_frame = page_object.mc_frame if page_object.mc_frame is not None else device_info.mc_frame
        if final_series and final_frame:
            cls._validate_series_frame(final_series, final_frame)

        # 自动填充 driver_code（若前端未传，根据品牌/系列/通信方式映射）
        await DriverService.fill_driver_code_for_device_vo(query_db, page_object)
        # 校验 driver_code 有效且驱动已启用
        await DriverService.validate_driver_code(query_db, page_object.driver_code)

        # 名称/编号长度校验（先于唯一性查询，避免超长值直达数据库）
        cls._validate_name_code_length(page_object.device_name, page_object.device_code)

        # 设备名称唯一性校验（排除自身）
        if page_object.device_name and page_object.device_name != (device_info.device_name or ''):
            existing_by_name = await DeviceDao.get_device_by_name(query_db, page_object.device_name)
            if existing_by_name:
                raise ServiceException(message=f'设备名称已存在：{page_object.device_name}')

        # 设备编号唯一性校验（排除自身）
        if page_object.device_code and page_object.device_code != (device_info.device_code or ''):
            existing_by_code = await DeviceDao.get_device_by_code(query_db, page_object.device_code)
            if existing_by_code:
                raise ServiceException(message=f'设备编号已存在：{page_object.device_code}')

        edit_device = page_object.model_dump(
            exclude_unset=True,
            exclude={'create_by', 'create_time', 'tags', 'del_flag'}  # 🔧 Day9/P1-17：编辑不可写 del_flag（防借 edit 权限软删）
        )

        try:
            device_name = device_info.device_name
            await DeviceDao.edit_device_dao(query_db, edit_device)
            await cls._log_change(
                query_db, 'update', 'device', page_object.id, device_name,
                change_by=page_object.update_by or '',
                before_value={'deviceName': device_info.device_name, 'deviceCode': device_info.device_code},
                after_value=edit_device
            )
            await query_db.commit()
            logger.info(f'编辑PLC设备成功：{device_name}')
            return CrudResponseModel(is_success=True, message='更新成功（请手动发布配置后采集节点方可生效）')
        except Exception as e:
            await query_db.rollback()
            logger.exception(f'编辑PLC设备失败：{e}')
            raise

    @classmethod
    async def set_device_status(
        cls, query_db: AsyncSession, device_id: int, status: str,
        update_by: str = '', update_time: datetime | None = None,
    ) -> None:
        """直接设置设备状态"""
        if status not in {'0', '1'}:
            raise ServiceException(message=f'无效的设备状态：{status}，仅支持 0（启用）或 1（停用）')
        device_info = await DeviceDao.get_device_detail_by_id(query_db, device_id)
        if not device_info:
            raise ServiceException(message=f'PLC设备不存在：ID={device_id}')
        update_dict: dict = {'id': device_id, 'status': status}
        if update_by:
            update_dict['update_by'] = update_by
        if update_time:
            update_dict['update_time'] = update_time
        await DeviceDao.edit_device_dao(query_db, update_dict)
        await cls._log_change(
            query_db, 'enable' if status == '0' else 'disable', 'device', device_id,
            device_info.device_name, change_by=update_by,
            before_value={'status': device_info.status},
            after_value={'status': status}
        )
        await query_db.commit()

    @classmethod
    async def disable_device_services(
        cls, query_db: AsyncSession, page_object: DeletePlcDeviceModel,
        update_by: str = '', update_time: datetime | None = None,
    ) -> CrudResponseModel:
        """
        停用PLC设备（设置status='1'，Node-RED不再采集但数据保留可见）

        :param query_db: orm对象
        :param page_object: 停用设备对象
        :param update_by: 操作者用户名
        :param update_time: 操作时间
        :return: 操作结果
        """
        if not page_object.ids:
            raise ServiceException(message='传入为空')

        try:
            id_list = [int(x) for x in page_object.ids.split(',') if x.strip()]
        except ValueError:
            raise ServiceException(message='ID列表格式不正确，必须为正整数')
        if not id_list:
            raise ServiceException(message='传入为空')

        try:
            # 批量查询所有目标设备（一次查询替代 N+1）
            result = await query_db.execute(
                sa_select(PlcDevice).where(
                    PlcDevice.id.in_(id_list),
                    PlcDevice.del_flag == '0',
                )
            )
            devices = result.scalars().all()
            existing_ids = {d.id for d in devices}

            # 批量更新状态
            if existing_ids:
                values = {'status': '1'}
                if update_by:
                    values['update_by'] = update_by
                if update_time:
                    values['update_time'] = update_time
                await query_db.execute(
                    sa_update(PlcDevice)
                    .where(PlcDevice.id.in_(list(existing_ids)))
                    .values(**values)
                )

                # 同步标记该设备通信状态为离线，避免 KPI 虚高
                await query_db.execute(
                    sa_update(DeviceCommStatus)
                    .where(DeviceCommStatus.device_id.in_(list(existing_ids)))
                    .values(online=0, error_msg='设备已停用', updated_at=datetime.now())
                )

                # 消除该设备所有未处理的 PLC_OFFLINE 告警（停用后 Node-RED 不再上报，告警不会自愈）
                for device_id in existing_ids:
                    await MonitorDao.resolve_alert(
                        query_db, 'PLC_OFFLINE', node_id=None, device_id=device_id
                    )

            # 记录日志
            for d in devices:
                logger.info(f'停用PLC设备成功：{d.device_name}')
                await cls._log_change(
                    query_db, 'disable', 'device', d.id, d.device_name,
                    change_by=update_by,
                    before_value={'status': '0'},
                    after_value={'status': '1'}
                )
            missing = set(id_list) - existing_ids
            for mid in missing:
                logger.warning(f'ID {mid} 对应的PLC设备不存在，跳过')

            await query_db.commit()
            return CrudResponseModel(is_success=True, message='操作成功（请手动发布配置后采集节点方可生效）')
        except Exception as e:
            await query_db.rollback()
            logger.exception(f'停用PLC设备失败：{e}')
            raise

    @classmethod
    async def soft_delete_device_services(
        cls, query_db: AsyncSession, page_object: DeletePlcDeviceModel,
        update_by: str = '', update_time: datetime | None = None,
    ) -> CrudResponseModel:
        """
        软删除PLC设备（设置del_flag='2'，同时软删除其下所有点位）

        :param query_db: orm对象
        :param page_object: 删除PLC设备对象
        :param update_by: 操作者用户名
        :param update_time: 操作时间
        :return: 删除结果
        """
        if not page_object.ids:
            raise ServiceException(message='传入为空')

        try:
            id_list = [int(x) for x in page_object.ids.split(',') if x.strip()]
        except ValueError:
            raise ServiceException(message='ID列表格式不正确，必须为正整数')
        if not id_list:
            raise ServiceException(message='传入为空')

        try:
            # 批量查询所有目标设备（一次查询替代 N+1）
            result = await query_db.execute(
                sa_select(PlcDevice).where(
                    PlcDevice.id.in_(id_list),
                    PlcDevice.del_flag == '0',
                )
            )
            devices = result.scalars().all()
            existing_ids = {d.id for d in devices}

            if existing_ids:
                # 构建审计字段
                audit_values = {}
                if update_by:
                    audit_values['update_by'] = update_by
                if update_time:
                    audit_values['update_time'] = update_time

                # 批量软删除该设备下的所有点位
                tag_values = {'del_flag': '2', **audit_values}
                await query_db.execute(
                    sa_update(PlcTag)
                    .where(
                        PlcTag.device_id.in_(list(existing_ids)),
                        PlcTag.del_flag == '0',
                    )
                    .values(**tag_values)
                )
                # 批量软删除设备
                device_values = {'del_flag': '2', **audit_values}
                await query_db.execute(
                    sa_update(PlcDevice)
                    .where(PlcDevice.id.in_(list(existing_ids)))
                    .values(**device_values)
                )
                # 同步清理设备通信状态，避免已删除设备仍被计入 KPI 在线数
                await query_db.execute(
                    sa_update(DeviceCommStatus)
                    .where(DeviceCommStatus.device_id.in_(list(existing_ids)))
                    .values(online=0, error_msg='设备已删除', updated_at=datetime.now())
                )

                # 消除该设备所有未处理的 PLC_OFFLINE 告警（删除后 Node-RED 不再上报，告警不会自愈）
                for device_id in existing_ids:
                    await MonitorDao.resolve_alert(
                        query_db, 'PLC_OFFLINE', node_id=None, device_id=device_id
                    )

                # 记录修改履历（软删除）
                for d in devices:
                    await cls._log_change(
                        query_db, 'delete', 'device', d.id, d.device_name, update_by,
                        before_value={'del_flag': '0'},
                        after_value={'del_flag': '2'},
                    )

            # 记录日志
            for d in devices:
                logger.info(f'删除PLC设备成功：{d.device_name}')
            missing = set(id_list) - existing_ids
            for mid in missing:
                logger.warning(f'ID {mid} 对应的PLC设备不存在，跳过')

            await query_db.commit()
            return CrudResponseModel(is_success=True, message='操作成功（请手动发布配置后采集节点方可生效）')
        except Exception as e:
            await query_db.rollback()
            logger.exception(f'删除PLC设备失败：{e}')
            raise

    @classmethod
    async def device_detail_services(cls, query_db: AsyncSession, device_id: int) -> PlcDeviceModel:
        """
        获取PLC设备详细信息（含点位列表）

        :param query_db: orm对象
        :param device_id: 设备ID
        :return: 设备详细信息
        """
        device = await DeviceDao.get_device_detail_by_id(query_db, device_id=device_id)
        if device:
            # 将ORM对象转为驼峰dict
            result_dict = CamelCaseUtil.transform_result(device)

            # 查询该设备下所有未删除的点位（is_page=False 时 paginate 不加分页限制，全量返回）
            tag_query = TagPageQueryModel()
            tag_result = await TagDao.get_tag_list_by_device_id(query_db, device_id, tag_query, is_page=False)
            result_dict['tags'] = [TagModel.model_validate(t) if isinstance(t, dict) else t for t in tag_result] if isinstance(tag_result, list) else []

            result = PlcDeviceModel.model_validate(result_dict)
        else:
            result = PlcDeviceModel()

        return result

    # ==================== 克隆 & 导出 ====================

    @classmethod
    async def clone_device_services(
        cls, query_db: AsyncSession, source_device_id: int,
        new_device_name: str, new_device_code: str, new_plc_ip: str,
        create_by: str = '',
    ) -> CrudResponseModel:
        """
        克隆设备：复制源设备所有字段 + 所有未删除点位

        :param query_db: orm对象
        :param source_device_id: 源设备ID
        :param new_device_name: 新设备名称
        :param new_device_code: 新设备编号
        :param new_plc_ip: 新设备IP
        :param create_by: 创建者用户名
        :return: 克隆结果
        """
        source = await DeviceDao.get_device_detail_by_id(query_db, source_device_id)
        if not source:
            raise ServiceException(message='源设备不存在')

        # 名称/编号长度校验（克隆建议值可能带叠加后缀，先于唯一性查询拦截超长值）
        cls._validate_name_code_length(new_device_name, new_device_code)

        # 校验新设备名称唯一性
        if new_device_name:
            existing_by_name = await DeviceDao.get_device_by_name(query_db, new_device_name)
            if existing_by_name:
                raise ServiceException(message=f'设备名称已存在：{new_device_name}')

        # 校验新设备编号唯一性
        if new_device_code:
            existing_by_code = await DeviceDao.get_device_by_code(query_db, new_device_code)
            if existing_by_code:
                raise ServiceException(message=f'设备编号已存在：{new_device_code}')

        if new_plc_ip:
            cls._validate_ip(new_plc_ip)
        # 校验新设备的通信方式是否需要 PLC IP
        cls._require_plc_ip_for_com_type(new_plc_ip or None, source.com_type)

        # 若源设备有 host_pc_ip 也做校验（克隆不要求改 host_pc_ip，但保留校验以防后续扩展）
        if source.host_pc_ip:
            cls._validate_ip(source.host_pc_ip, allow_port=True)
        if source.backup_pc_ip:
            cls._validate_ip(source.backup_pc_ip, allow_port=True)

        now = datetime.now()
        try:
            # 1. 复制设备（排除 id/审计字段），覆盖 name/code/ip
            source_dict = CamelCaseUtil.transform_result(source)
            exclude_keys = {'id', 'createBy', 'createTime', 'updateBy', 'updateTime', 'tags', 'tagCount'}
            new_device_dict = {k: v for k, v in source_dict.items() if k not in exclude_keys}
            new_device_dict['deviceName'] = new_device_name
            new_device_dict['deviceCode'] = new_device_code
            new_device_dict['plcIp'] = new_plc_ip
            new_device_dict['createBy'] = create_by
            new_device_dict['createTime'] = now
            new_device_dict['updateBy'] = create_by
            new_device_dict['updateTime'] = now

            new_device_vo = PlcDeviceModel.model_validate(new_device_dict)
            db_new_device = await DeviceDao.add_device_dao(query_db, new_device_vo)
            await query_db.flush()

            # 2. 复制点位（is_page=False 时 paginate 不加分页限制，全量复制所有未删除点位）
            tag_query = TagPageQueryModel()
            source_tags = await TagDao.get_tag_list_by_device_id(
                query_db, source_device_id, tag_query, is_page=False
            )
            if isinstance(source_tags, list):
                for tag_dict in source_tags:
                    tag_data = {k: v for k, v in (tag_dict if isinstance(tag_dict, dict) else {}).items()
                                if k not in ['id', 'createBy', 'createTime', 'updateBy', 'updateTime']}
                    tag_data['deviceId'] = db_new_device.id
                    tag_data['createBy'] = create_by
                    tag_data['createTime'] = now
                    tag_data['updateBy'] = create_by
                    tag_data['updateTime'] = now
                    tag_vo = TagModel.model_validate(tag_data)
                    await TagDao.add_tag_dao(query_db, tag_vo)

            source_name = source.device_name  # commit 前保存，避免 MissingGreenlet
            await query_db.commit()
            logger.info(f'克隆PLC设备成功：{source_name} → {new_device_name}')
            return CrudResponseModel(is_success=True, message=f'克隆成功：{source_name} → {new_device_name}（请手动发布配置后采集节点方可生效）')
        except Exception as e:
            await query_db.rollback()
            logger.exception(f'克隆PLC设备失败：{e}')
            raise

    @classmethod
    async def export_device_list_services(cls, query_db: AsyncSession, query_object: PlcDevicePageQueryModel) -> bytes:
        """
        导出设备列表（含点位）为 Excel

        :param query_db: orm对象
        :param query_object: 查询参数
        :return: Excel 二进制数据
        """
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        # 使用 write_only 模式 + 分页写入，避免大数据量 OOM
        wb = Workbook(write_only=True)
        ws = wb.create_sheet('PLC设备清单')

        device_headers = [
            '设备编号', '设备名称', 'PLC品牌', 'PLC系列', '通信方式',
            'PLC IP', 'PLC端口', '采集PC IP', '备采集PC IP', 'MES IP', 'MES端口', '帧格式',
            '站号', '网络号', '采集周期(ms)', '通信超时(ms)', '重试次数', '重试间隔(ms)',
            '触发方式', '状态', '备注'
        ]
        device_keys = [
            'deviceCode', 'deviceName', 'plcBrand', 'plcSeries', 'comType',
            'plcIp', 'plcPort', 'hostPcIp', 'backupPcIp', 'mesIp', 'mesPort', 'mcFrame',
            'stationNo', 'networkNo', 'scanIntervalMs', 'commTimeoutMs', 'retryCount', 'retryIntervalMs',
            'triggerKind', 'status', 'remark'
        ]

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        # 分页流式写入：每批500条
        page_size = 500
        page_num = 1
        row_idx = 0

        while True:
            query_object.page_num = page_num
            query_object.page_size = page_size
            device_result = await DeviceDao.get_device_list(query_db, query_object, is_page=True)
            page_data = device_result.rows if hasattr(device_result, 'rows') else []

            if not page_data:
                break

            for dev in page_data:
                row_data = [dev.get(key, '') if isinstance(dev, dict) else '' for key in device_keys]
                # 表头样式只对第一行有效，write_only 模式下需单独处理
                if row_idx == 0:
                    header_row = [h for h in device_headers]
                    ws.append(header_row)
                    row_idx += 1
                ws.append(row_data)
                row_idx += 1

            if len(page_data) < page_size:
                break
            page_num += 1

        binary_data = io.BytesIO()
        wb.save(binary_data)
        binary_data.seek(0)
        return binary_data.getvalue()
