"""PLC 采集点位服务层"""
import json
from datetime import datetime
from typing import Any, Union

from sqlalchemy import func as sa_func, or_ as sa_or, select as sa_select, update as sa_update
from sqlalchemy.ext.asyncio import AsyncSession

from common.vo import CrudResponseModel, PageModel
from exceptions.exception import ServiceException
from module_plc.dao.device_dao import DeviceDao
from module_plc.dao.tag_dao import TagDao
from module_plc.entity.do.change_log_do import PlcChangeLog
from module_plc.entity.do.monitor_do import DeviceCommStatus
from module_plc.entity.do.tag_do import PlcTag
from module_plc.entity.vo.tag_vo import (
    TagBatchUpdateModel,
    TagImportError,
    TagImportResult,
    TagModel,
    TagPageQueryModel,
)
from utils.common_util import CamelCaseUtil
from utils.log_util import logger

# ==================== 共享校验 ====================

VALID_REGISTER_TYPES = {'D', 'W', 'X', 'Y', 'M', 'R', 'L', 'B', 'HR', 'IR', 'CR', 'COIL', 'DI', 'DISCRETE', 'DB', 'I', 'Q'}  # 🔧 Day5: 补 DISCRETE；🔧 2026-08-23: 补 S7 区域 DB/I/Q（计算点位走 is_calc 分支不进此校验）
VALID_DATA_TYPES = {'INT16', 'INT32', 'FLOAT', 'BIT', 'UINT16', 'UINT32', 'BOOL', 'DOUBLE'}
VALID_CALC_OPS = ('sum', 'avg', 'min', 'max')  # 计算点位算子（calc_op 非空即为计算点位，寄存器字段存占位 CALC）


def _validate_register_type(register_type: str) -> None:
    if register_type and register_type.upper() not in VALID_REGISTER_TYPES:
        raise ServiceException(
            message=f'无效的寄存器类型：{register_type}，有效值为：{", ".join(sorted(VALID_REGISTER_TYPES))}'
        )


def _validate_data_type(data_type: str) -> None:
    if data_type and data_type.upper() not in VALID_DATA_TYPES:
        raise ServiceException(
            message=f'无效的数据类型：{data_type}，有效值为：{", ".join(sorted(VALID_DATA_TYPES))}'
        )


def _validate_address_format(register_type: str, register_address: str, plc_series: str = '', driver_code: str = '') -> None:
    """🔧 Day9/P1-19：寄存器地址格式校验（按软元件进制规则，与边缘 parseDeviceAddress 一致）

    W/B 及非 FX 系列的 X/Y 为 16 进制；FX 系列 X/Y 为 8 进制；其余 10 进制。
    🔧 2026-08-23：S7 驱动（siemens_s7）按 TIA 风格地址校验（DB1.DBW0/MW10/I0.0），与 s7 包地址解析一致。
    """
    import re
    addr = str(register_address or '').strip()
    if not addr:
        raise ServiceException(message='寄存器地址不能为空')
    if driver_code == 'siemens_s7':
        # TIA 风格：DB 区 DB<n>.DBW/D/X<字节>[.位]；M/I/Q 区 [W/D 前缀]字节[.位]
        if not re.fullmatch(r'(DB[0-9]+\.DB[WDX][0-9]+(\.[0-7])?|[MIQ][WD]?[0-9]+(\.[0-7])?)', addr, re.IGNORECASE):
            raise ServiceException(message=f'S7 地址格式错误：{addr}（示例：DB1.DBW0 / DB1.DBX0.0 / MW10 / I0.0）')
        return
    rt = (register_type or '').upper()
    series = (plc_series or '').upper().replace('-', '')
    is_fx = series.startswith('FX') or series.startswith('IQF')
    if rt in ('W', 'B') or (rt in ('X', 'Y') and not is_fx):
        if not re.fullmatch(r'[0-9A-Fa-f]+', addr):
            raise ServiceException(message=f'寄存器地址格式错误：{rt}{addr}（{rt} 为 16 进制地址）')
    elif rt in ('X', 'Y') and is_fx:
        if not re.fullmatch(r'[0-7]+', addr):
            raise ServiceException(message=f'寄存器地址格式错误：{rt}{addr}（FX 系列 X/Y 为 8 进制地址）')
    else:
        if not re.fullmatch(r'[0-9]+', addr):
            raise ServiceException(message=f'寄存器地址格式错误：{rt}{addr}（{rt} 为 10 进制地址）')


def _safe_float(val, default: float = 0.0) -> float:
    """安全浮点转换，避免 0 被默认值覆盖，避免 '1000.0' 类字符串失败"""
    if val is None or str(val).strip() == '':
        return default
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return default


def _safe_int(val, default: int = 0) -> int:
    """安全整数转换，先转 float 再转 int，兼容 '1000.0'"""
    if val is None or str(val).strip() == '':
        return default
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return default


# ==================== TagService ====================

class TagService:
    """PLC采集点位服务层"""

    @classmethod
    async def _sync_device_comm_status(cls, query_db: AsyncSession, device_id: int) -> None:
        """点位停用/删除后同步设备通信状态：若设备下无启用且未删除点位，则标记离线。"""
        if not device_id:
            return
        # 统计设备下仍启用且未删除的点位数
        result = await query_db.execute(
            sa_select(sa_func.count(PlcTag.id)).where(
                PlcTag.device_id == device_id,
                PlcTag.del_flag == '0',
                PlcTag.status == '0',
            )
        )
        enabled_count = result.scalar() or 0
        if enabled_count == 0:
            await query_db.execute(
                sa_update(DeviceCommStatus)
                .where(DeviceCommStatus.device_id == device_id)
                .values(online=0, error_msg='设备下无启用点位', updated_at=datetime.now())
            )
            logger.info(f'设备 {device_id} 下无启用点位，通信状态已标记为离线')

    @staticmethod
    def _json_safe(value: dict | None) -> dict | None:
        """将 dict 转为 JSON 可序列化形式（datetime 等类型转字符串），避免 JSON 列写入失败"""
        if value is None:
            return None
        try:
            return json.loads(json.dumps(value, ensure_ascii=False, default=str))
        except (TypeError, ValueError):
            return {'_raw': str(value)}

    @classmethod
    async def _log_change(
        cls, db: AsyncSession, change_type: str, target_type: str,
        target_id: int, target_name: str = '', change_by: str = '',
        before_value: dict | None = None, after_value: dict | None = None,
        remark: str | None = None,
    ) -> None:
        """记录修改履历（失败只告警，不阻塞业务；字段截断到列长，防止 commit 时 flush 失败毒化主事务）"""
        try:
            log = PlcChangeLog(
                change_type=change_type[:20],
                target_type=target_type[:20],
                target_id=target_id,
                target_name=(target_name or '')[:100],
                change_by=(change_by or '')[:64],
                before_value=cls._json_safe(before_value),
                after_value=cls._json_safe(after_value),
                remark=(remark or '')[:500] if remark else None,
            )
            db.add(log)
        except Exception as exc:
            logger.warning(f'修改履历记录失败: {exc}')

    @classmethod
    async def _validate_calc_tag(cls, query_db: AsyncSession, page_object: TagModel, exclude_tag_id: int | None = None) -> None:
        """计算点位校验：算子合法 + 来源同设备/存在/非计算/非自身 + 所在设备至少有 1 个启用采集点位。

        计算点位不参与驱动读取（寄存器字段为占位 CALC），由边缘数据管道在采集轮次中求值。
        """
        op = (page_object.calc_op or '').strip().lower()
        if op not in VALID_CALC_OPS:
            raise ServiceException(message=f'计算方式不合法：{page_object.calc_op}（仅支持 {"、".join(VALID_CALC_OPS)}）')
        page_object.calc_op = op
        raw = page_object.calc_source_ids
        if isinstance(raw, str):
            try:
                raw = json.loads(raw)
            except (TypeError, ValueError):
                raise ServiceException(message='来源点位格式错误（应为ID数组）')
        source_ids = []
        for x in raw or []:
            try:
                source_ids.append(int(x))
            except (TypeError, ValueError):
                raise ServiceException(message=f'来源点位ID非法：{x}')
        source_ids = list(dict.fromkeys(source_ids))
        if not source_ids:
            raise ServiceException(message='计算点位必须选择至少 1 个来源点位')
        if exclude_tag_id and exclude_tag_id in source_ids:
            raise ServiceException(message='来源点位不能包含计算点位自身')
        rows = (await query_db.execute(
            sa_select(PlcTag.id, PlcTag.device_id, PlcTag.calc_op).where(
                PlcTag.id.in_(source_ids), PlcTag.del_flag == '0',
            )
        )).all()
        found = {r.id: r for r in rows}
        missing = [sid for sid in source_ids if sid not in found]
        if missing:
            raise ServiceException(message=f'来源点位不存在或已删除：{missing}')
        for r in rows:
            if r.calc_op:
                raise ServiceException(message=f'来源点位ID={r.id} 是计算点位，禁止链式引用')
            if r.device_id != page_object.device_id:
                raise ServiceException(message=f'来源点位ID={r.id} 与计算点位不在同一设备（v1 仅限同设备）')
        real_cnt = (await query_db.execute(
            sa_select(sa_func.count(PlcTag.id)).where(
                PlcTag.device_id == page_object.device_id,
                PlcTag.del_flag == '0',
                PlcTag.status == '0',
                sa_or(PlcTag.calc_op.is_(None), PlcTag.calc_op == ''),
                PlcTag.id != (exclude_tag_id or 0),
            )
        )).scalar() or 0
        if real_cnt == 0:
            raise ServiceException(message='计算点位所在设备需至少保留 1 个启用中的采集点位（计算由采集轮次驱动）')
        # DB 列为 VARCHAR，统一落 JSON 文本（消费侧：前端/边缘 cm 各自 JSON.parse）
        page_object.calc_source_ids = json.dumps(source_ids, ensure_ascii=False)

    @classmethod
    async def _assert_not_calc_source(cls, query_db: AsyncSession, id_list: list[int], action: str = '删除') -> None:
        """计算点位引用守卫：被启用中计算点位引用的点位不可删除/停用（否则计算点位断源）。"""
        rows = (await query_db.execute(
            sa_select(PlcTag.id, PlcTag.tag_name, PlcTag.calc_source_ids).where(
                PlcTag.del_flag == '0',
                PlcTag.status == '0',
                PlcTag.calc_op.isnot(None),
                PlcTag.calc_op != '',
            )
        )).all()
        targets = set(id_list)
        for r in rows:
            try:
                src = json.loads(r.calc_source_ids) if isinstance(r.calc_source_ids, str) else (r.calc_source_ids or [])
            except (TypeError, ValueError):
                src = []
            if targets & set(src):
                raise ServiceException(
                    message=f'点位被计算点位「{r.tag_name}」(ID={r.id}) 引用，不可{action}；请先调整该计算点位的来源'
                )

    @classmethod
    async def add_tag_services(cls, query_db: AsyncSession, page_object: TagModel) -> CrudResponseModel:
        device_info = None
        if page_object.device_id is not None:
            device_info = await DeviceDao.get_device_detail_by_id(query_db, page_object.device_id)
            if not device_info:
                raise ServiceException(message=f'所属PLC设备不存在：ID={page_object.device_id}')
        # 计算点位：寄存器字段存占位 CALC，跳过寄存器/地址校验；来源与设备约束单独校验
        is_calc = bool(page_object.calc_op and str(page_object.calc_op).strip())
        if is_calc:
            await cls._validate_calc_tag(query_db, page_object)
            page_object.register_type = 'CALC'
            page_object.register_address = 'CALC'
            if not page_object.data_type:
                page_object.data_type = 'DOUBLE'
        if not is_calc and not page_object.register_type:
            raise ServiceException(message='寄存器类型不能为空')
        if not page_object.data_type:
            raise ServiceException(message='数据类型不能为空')
        if not is_calc:
            _validate_register_type(page_object.register_type)
            _validate_data_type(page_object.data_type)
            # 🔧 Day9/P1-19：地址格式按设备系列的进制规则校验（S7 驱动按 TIA 风格，由 driver_code 区分）
            _validate_address_format(
                page_object.register_type, page_object.register_address,
                (device_info.plc_series if device_info else '') or '',
                (device_info.driver_code if device_info else '') or '',
            )

        # 🔧 Day9/P1-16：同设备下点位名/寄存器地址查重（DB 无唯一约束，应用层兜底，导入路径已有判重）
        dup_name = (await query_db.execute(
            sa_select(PlcTag.id).where(
                PlcTag.device_id == page_object.device_id,
                PlcTag.tag_name == page_object.tag_name,
                PlcTag.del_flag == '0',
            )
        )).first()
        if dup_name:
            raise ServiceException(message=f'点位名称在该设备下已存在：{page_object.tag_name}')
        # 计算点位不查地址重复（多个计算点位占位地址同为 CALC/CALC，本就不参与读取）
        if not is_calc:
            dup_addr = (await query_db.execute(
                sa_select(PlcTag.id).where(
                    PlcTag.device_id == page_object.device_id,
                    PlcTag.register_type == page_object.register_type,
                    PlcTag.register_address == page_object.register_address,
                    PlcTag.del_flag == '0',
                )
            )).first()
            if dup_addr:
                raise ServiceException(message=f'寄存器地址在该设备下已被占用：{page_object.register_type}{page_object.register_address}')

        try:
            db_tag = await TagDao.add_tag_dao(query_db, page_object)
            tag_name = db_tag.tag_name
            await cls._log_change(
                query_db, 'add', 'tag', db_tag.id, tag_name,
                change_by=page_object.create_by or page_object.update_by or '',
                after_value={
                    'tagName': tag_name,
                    'registerType': db_tag.register_type,
                    'registerAddress': db_tag.register_address,
                    'dataType': db_tag.data_type,
                    'deviceId': db_tag.device_id,
                },
            )
            await query_db.commit()
            logger.info(f'新增PLC点位成功：{tag_name}')
            return CrudResponseModel(is_success=True, message='新增成功（请手动发布配置后采集节点方可生效）')
        except Exception as e:
            await query_db.rollback()
            logger.exception(f'新增PLC点位失败：{e}')
            raise

    @classmethod
    async def edit_tag_services(cls, query_db: AsyncSession, page_object: TagModel) -> CrudResponseModel:
        tag_info = await TagDao.get_tag_detail_by_id(query_db, page_object.id)
        if not tag_info:
            raise ServiceException(message='PLC点位不存在')

        if page_object.device_id is not None:
            device_info = await DeviceDao.get_device_detail_by_id(query_db, page_object.device_id)
            if not device_info:
                raise ServiceException(message=f'所属PLC设备不存在：ID={page_object.device_id}')

        # 计算点位：calc_op 未提交则沿用库内值；寄存器字段为占位 CALC，跳过寄存器/地址校验
        effective_calc_op = page_object.calc_op if page_object.calc_op is not None else tag_info.calc_op
        is_calc = bool(effective_calc_op and str(effective_calc_op).strip())
        if is_calc:
            page_object.calc_op = str(effective_calc_op).strip().lower()
            if page_object.device_id is None:
                page_object.device_id = tag_info.device_id
            if page_object.calc_source_ids is None:
                page_object.calc_source_ids = tag_info.calc_source_ids
            await cls._validate_calc_tag(query_db, page_object, exclude_tag_id=tag_info.id)
            page_object.register_type = 'CALC'
            page_object.register_address = 'CALC'

        if not is_calc:
            if page_object.register_type is not None:
                _validate_register_type(page_object.register_type)
            if page_object.data_type is not None:
                _validate_data_type(page_object.data_type)
            # 🔧 Day9/P1-19：编辑同样做地址进制校验（S7 驱动按 TIA 风格，由 driver_code 区分）
            if page_object.register_address is not None:
                _validate_address_format(
                    page_object.register_type or tag_info.register_type,
                    page_object.register_address,
                    (device_info.plc_series if device_info else '') or '',
                    (device_info.driver_code if device_info else (tag_info.device.driver_code if getattr(tag_info, 'device', None) else '')) or '',
                )

        # 🔧 Day9/P1-16：编辑同样查重（排除自身）
        # 🔧 2026-08-22：查重范围只看「启用中」的兄弟点位——停用点位不参与采集，
        #    不应阻塞在役点位的编辑（此前停用同名点位会导致编辑报"名称已存在"）
        if page_object.tag_name is not None:
            dup_name = (await query_db.execute(
                sa_select(PlcTag.id).where(
                    PlcTag.device_id == (page_object.device_id or tag_info.device_id),
                    PlcTag.tag_name == page_object.tag_name,
                    PlcTag.del_flag == '0',
                    PlcTag.status == '0',
                    PlcTag.id != page_object.id,
                )
            )).first()
            if dup_name:
                raise ServiceException(message=f'点位名称在该设备下已存在：{page_object.tag_name}')
        if not is_calc and page_object.register_type is not None and page_object.register_address is not None:
            dup_addr = (await query_db.execute(
                sa_select(PlcTag.id).where(
                    PlcTag.device_id == (page_object.device_id or tag_info.device_id),
                    PlcTag.register_type == page_object.register_type,
                    PlcTag.register_address == page_object.register_address,
                    PlcTag.del_flag == '0',
                    PlcTag.status == '0',
                    PlcTag.id != page_object.id,
                )
            )).first()
            if dup_addr:
                raise ServiceException(message=f'寄存器地址在该设备下已被占用：{page_object.register_type}{page_object.register_address}')

        edit_tag = page_object.model_dump(exclude_unset=True, exclude={'device_name', 'del_flag'})  # 🔧 Day9/P1-17：编辑不可写 del_flag（防借 edit 权限软删）
        try:
            tag_name = tag_info.tag_name
            await TagDao.edit_tag_dao(query_db, edit_tag)
            await cls._log_change(
                query_db, 'update', 'tag', page_object.id, tag_name,
                change_by=page_object.update_by or '',
                before_value={
                    'tagName': tag_info.tag_name,
                    'registerType': tag_info.register_type,
                    'registerAddress': tag_info.register_address,
                    'dataType': tag_info.data_type,
                },
                after_value=edit_tag,
            )
            await query_db.commit()
            logger.info(f'编辑PLC点位成功：{tag_name}')
            return CrudResponseModel(is_success=True, message='更新成功（请手动发布配置后采集节点方可生效）')
        except Exception as e:
            await query_db.rollback()
            logger.exception(f'编辑PLC点位失败：{e}')
            raise

    @classmethod
    async def soft_delete_tag_services(
        cls, query_db: AsyncSession, page_object,
        update_by: str = '', update_time=None,
    ) -> CrudResponseModel:
        if not page_object.ids:
            raise ServiceException(message='传入为空')
        try:
            id_list = [int(x) for x in page_object.ids.split(',') if x.strip()]
        except ValueError:
            raise ServiceException(message='ID列表格式不正确，必须为正整数')
        if not id_list:
            raise ServiceException(message='传入为空')
        await cls._assert_not_calc_source(query_db, id_list, action='删除')

        try:
            # 批量查询所有目标点位（一次查询替代 N+1）
            result = await query_db.execute(
                sa_select(PlcTag).where(
                    PlcTag.id.in_(id_list),
                    PlcTag.del_flag == '0',
                )
            )
            tags = result.scalars().all()
            existing_ids = {t.id for t in tags}

            # 批量软删除
            if existing_ids:
                values = {'del_flag': '2'}
                if update_by:
                    values['update_by'] = update_by
                if update_time:
                    values['update_time'] = update_time
                await query_db.execute(
                    sa_update(PlcTag)
                    .where(PlcTag.id.in_(list(existing_ids)))
                    .values(**values)
                )
                # 记录修改履历（软删除）
                for t in tags:
                    await cls._log_change(
                        query_db, 'delete', 'tag', t.id, t.tag_name, update_by,
                        before_value={'del_flag': '0'},
                        after_value={'del_flag': '2'},
                    )

            # 记录日志
            for t in tags:
                logger.info(f'删除PLC点位成功：{t.tag_name}')
            missing = set(id_list) - existing_ids
            for mid in missing:
                logger.warning(f'ID {mid} 对应的PLC点位不存在，跳过')

            # 同步设备通信状态：若设备下无启用点位则标记离线
            device_ids = {t.device_id for t in tags if t.device_id}
            for device_id in device_ids:
                await cls._sync_device_comm_status(query_db, device_id)

            await query_db.commit()
            return CrudResponseModel(is_success=True, message='删除成功（请手动发布配置后采集节点方可生效）')
        except Exception as e:
            await query_db.rollback()
            logger.exception(f'删除PLC点位失败：{e}')
            raise

    @classmethod
    async def disable_tag_services(
        cls, query_db: AsyncSession, page_object,
        update_by: str = '', update_time=None,
    ) -> CrudResponseModel:
        if not page_object.ids:
            raise ServiceException(message='传入为空')
        try:
            id_list = [int(x) for x in page_object.ids.split(',') if x.strip()]
        except ValueError:
            raise ServiceException(message='ID列表格式不正确，必须为正整数')
        if not id_list:
            raise ServiceException(message='传入为空')
        await cls._assert_not_calc_source(query_db, id_list, action='停用')

        try:
            # 批量查询所有目标点位（一次查询替代 N+1）
            result = await query_db.execute(
                sa_select(PlcTag).where(
                    PlcTag.id.in_(id_list),
                    PlcTag.del_flag == '0',
                )
            )
            tags = result.scalars().all()
            existing_ids = {t.id for t in tags}

            # 批量更新状态
            if existing_ids:
                values = {'status': '1'}
                if update_by:
                    values['update_by'] = update_by
                if update_time:
                    values['update_time'] = update_time
                await query_db.execute(
                    sa_update(PlcTag)
                    .where(PlcTag.id.in_(list(existing_ids)))
                    .values(**values)
                )
                # 记录修改履历（停用）
                for t in tags:
                    await cls._log_change(
                        query_db, 'disable', 'tag', t.id, t.tag_name, update_by,
                        before_value={'status': t.status},
                        after_value={'status': '1'},
                    )

            # 记录日志
            for t in tags:
                logger.info(f'停用PLC点位成功：{t.tag_name}')
            missing = set(id_list) - existing_ids
            for mid in missing:
                logger.warning(f'ID {mid} 对应的PLC点位不存在，跳过')

            # 同步设备通信状态：若设备下无启用点位则标记离线
            device_ids = {t.device_id for t in tags if t.device_id}
            for device_id in device_ids:
                await cls._sync_device_comm_status(query_db, device_id)

            await query_db.commit()
            return CrudResponseModel(is_success=True, message='操作成功（请手动发布配置后采集节点方可生效）')
        except Exception as e:
            await query_db.rollback()
            logger.exception(f'停用PLC点位失败：{e}')
            raise

    @classmethod
    async def tag_detail_services(cls, query_db: AsyncSession, tag_id: int) -> TagModel:
        tag = await TagDao.get_tag_detail_by_id(query_db, tag_id=tag_id)
        return TagModel.model_validate(CamelCaseUtil.transform_result(tag)) if tag else TagModel()

    @classmethod
    async def get_tag_list_services(
        cls, query_db: AsyncSession, device_id: int, query_object, is_page: bool = False
    ) -> Union[PageModel, list[dict[str, Any]]]:
        return await TagDao.get_tag_list_by_device_id(query_db, device_id, query_object, is_page)

    @classmethod
    async def get_tag_global_list_services(
        cls, query_db: AsyncSession, query_object, is_page: bool = False
    ) -> Union[PageModel, list[dict[str, Any]]]:
        return await TagDao.get_tag_global_list(query_db, query_object, is_page)

    @classmethod
    async def batch_update_tag_services(
        cls, query_db: AsyncSession, batch_data: TagBatchUpdateModel,
        update_by: str = '', update_time=None,
    ) -> CrudResponseModel:
        if not batch_data.ids:
            raise ServiceException(message='传入为空')
        try:
            id_list = [int(x) for x in batch_data.ids.split(',') if x.strip()]
        except ValueError:
            raise ServiceException(message='ID列表格式不正确，必须为正整数')
        update_dict = {}
        if batch_data.register_type:
            _validate_register_type(batch_data.register_type)
            update_dict['register_type'] = batch_data.register_type.upper()
        if batch_data.data_type:
            _validate_data_type(batch_data.data_type)
            update_dict['data_type'] = batch_data.data_type.upper()
        if batch_data.status is not None:
            update_dict['status'] = batch_data.status
        if batch_data.unit is not None:
            update_dict['unit'] = batch_data.unit
        if batch_data.transform_type is not None:
            update_dict['transform_type'] = batch_data.transform_type
        if batch_data.transform_slope_a is not None:
            update_dict['transform_slope_a'] = batch_data.transform_slope_a
        if batch_data.transform_offset_b is not None:
            update_dict['transform_offset_b'] = batch_data.transform_offset_b
        if batch_data.eng_unit is not None:
            update_dict['eng_unit'] = batch_data.eng_unit
        if batch_data.report_deadband_ms is not None:
            update_dict['report_deadband_ms'] = batch_data.report_deadband_ms
        if batch_data.report_force_interval_ms is not None:
            update_dict['report_force_interval_ms'] = batch_data.report_force_interval_ms
        # 批量设置协议参数
        if batch_data.bit_offset is not None:
            update_dict['bit_offset'] = batch_data.bit_offset
        if batch_data.byte_order is not None:
            update_dict['byte_order'] = batch_data.byte_order
        if batch_data.word_order is not None:
            update_dict['word_order'] = batch_data.word_order
        if batch_data.protocol_params is not None:
            update_dict['protocol_params'] = batch_data.protocol_params
        if not update_dict:
            raise ServiceException(message='没有需要更新的字段')
        if update_by:
            update_dict['update_by'] = update_by
        if update_time:
            update_dict['update_time'] = update_time
        try:
            await TagDao.batch_update_tags(query_db, id_list, update_dict)
            await cls._log_change(
                query_db, 'update', 'tag', 0, f'批量更新{len(id_list)}条点位', update_by,
                after_value={k: v for k, v in update_dict.items() if k not in ('update_by', 'update_time')},
                remark=f'点位ID列表: {batch_data.ids[:450]}',
            )
            await query_db.commit()
            logger.info(f'批量更新点位成功：{len(id_list)} 条')
            return CrudResponseModel(is_success=True, message=f'成功更新 {len(id_list)} 条点位（请手动发布配置后采集节点方可生效）')
        except Exception as e:
            await query_db.rollback()
            logger.exception(f'批量更新点位失败：{e}')
            raise

    @staticmethod
    def parse_import_file(content: bytes, filename: str) -> list[dict]:
        """
        解析导入文件（Excel/JSON），返回点位字典列表

        :param content: 文件二进制内容
        :param filename: 文件名（用于判断格式）
        :return: 点位字典列表
        :raises ServiceException: 文件中未解析到任何数据时抛出
        """
        import io
        import json
        from openpyxl import load_workbook

        if filename and filename.lower().endswith('.json'):
            try:
                data = json.loads(content.decode('utf-8'))
            except (json.JSONDecodeError, UnicodeDecodeError) as e:
                raise ServiceException(message=f'JSON文件解析失败：{e}')
            return data if isinstance(data, list) else data.get('tags', data.get('data', []))

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise ServiceException(message='文件中未解析到任何数据')
            headers = [str(h).strip() if h else '' for h in rows[0]]
            header_map = {
                '点位名称': 'tagName', '寄存器类型': 'registerType', '寄存器地址': 'registerAddress',
                '数据类型': 'dataType', '单位': 'unit', '描述': 'description', '排序号': 'sortOrder',
                '换算类型': 'transformType', '斜率/乘数a': 'transformSlopeA', '偏移量b': 'transformOffsetB',
                '工程单位': 'engUnit', '死区(数值)': 'reportDeadbandMs',
                '强制间隔(ms)': 'reportForceIntervalMs',
                '位偏移': 'bitOffset', '字节序': 'byteOrder', '字序': 'wordOrder',
                'tag_name': 'tagName', 'register_type': 'registerType', 'register_address': 'registerAddress',
                'data_type': 'dataType', 'sort_order': 'sortOrder',
                'transform_type': 'transformType', 'transform_slope_a': 'transformSlopeA',
                'transform_offset_b': 'transformOffsetB', 'eng_unit': 'engUnit',
                'report_deadband_ms': 'reportDeadbandMs', 'report_force_interval_ms': 'reportForceIntervalMs',
                'bit_offset': 'bitOffset', 'byte_order': 'byteOrder', 'word_order': 'wordOrder',
            }
            tag_list = []
            for row_data in rows[1:]:
                if not any(row_data):
                    continue
                item = {}
                for idx, h in enumerate(headers):
                    key = header_map.get(h, h)
                    val = row_data[idx] if idx < len(row_data) else None
                    item[key] = str(val).strip() if val is not None else ''
                tag_list.append(item)
            return tag_list
        finally:
            wb.close()

    @classmethod
    async def import_tag_services(
        cls, query_db: AsyncSession, device_id: int, tag_list: list[dict],
        create_by: str = '',
    ) -> TagImportResult:
        device_info = await DeviceDao.get_device_detail_by_id(query_db, device_id)
        if not device_info:
            raise ServiceException(message=f'所属PLC设备不存在：ID={device_id}')

        # 查询设备下已存在的点位名称（用于同设备内重名校验，防止脏数据或批量 INSERT 整体失败）
        existing_result = await query_db.execute(
            sa_select(PlcTag.tag_name).where(
                PlcTag.device_id == device_id,
                PlcTag.del_flag == '0',
            )
        )
        existing_names = {row[0] for row in existing_result}
        seen_names: set[str] = set()

        now = datetime.now()
        success_count = 0
        errors = []
        valid_tags: list[TagModel] = []
        for i, item in enumerate(tag_list):
            row_num = i + 2
            try:
                tag_name = str(item.get('tagName', item.get('tag_name', item.get('点位名称', '')))).strip()
                register_type = str(item.get('registerType', item.get('register_type', item.get('寄存器类型', '')))).strip()
                register_address = str(item.get('registerAddress', item.get('register_address', item.get('寄存器地址', '')))).strip()
                data_type = str(item.get('dataType', item.get('data_type', item.get('数据类型', '')))).strip()

                if not tag_name or not register_type or not register_address or not data_type:
                    missing = []
                    if not tag_name: missing.append('点位名称')
                    if not register_type: missing.append('寄存器类型')
                    if not register_address: missing.append('寄存器地址')
                    if not data_type: missing.append('数据类型')
                    errors.append(TagImportError(row=row_num, reason=f'必填字段缺失：{", ".join(missing)}'))
                    continue

                # 同设备内重名校验：先查 DB 已存在，再查文件内重复
                if tag_name in existing_names:
                    errors.append(TagImportError(row=row_num, reason=f'点位名称 "{tag_name}" 在该设备中已存在'))
                    continue
                if tag_name in seen_names:
                    errors.append(TagImportError(row=row_num, reason=f'点位名称 "{tag_name}" 在文件内重复'))
                    continue
                seen_names.add(tag_name)

                if register_type.upper() not in VALID_REGISTER_TYPES:
                    errors.append(TagImportError(row=row_num, reason=f'无效的寄存器类型 "{register_type}"'))
                    continue
                if data_type.upper() not in VALID_DATA_TYPES:
                    errors.append(TagImportError(row=row_num, reason=f'无效的数据类型 "{data_type}"'))
                    continue

                unit = str(item.get('unit', '') or item.get('单位', '')).strip()
                description = str(item.get('description', '') or item.get('描述', '')).strip()
                sort_str = str(item.get('sortOrder', item.get('sort_order', item.get('排序号', 0)))).strip()
                # 支持 '1.5' 这类浮点字符串先转 float 再取整，避免 isdigit 失败时静默归零
                sort_order = int(float(sort_str)) if sort_str else 0

                # 读取可选字段（使用安全转换，避免0被覆盖、避免'1000.0'解析失败）
                transform_type = str(item.get('transformType', 'none')).strip() or 'none'
                transform_slope_a = _safe_float(item.get('transformSlopeA'), 1.0)
                transform_offset_b = _safe_float(item.get('transformOffsetB'), 0.0)
                eng_unit = str(item.get('engUnit', '')).strip() or None
                report_deadband_ms = _safe_int(item.get('reportDeadbandMs'), 1000)
                report_force_interval_ms = _safe_int(item.get('reportForceIntervalMs'), 5000)
                bit_offset = _safe_int(item.get('bitOffset'), None)
                byte_order = str(item.get('byteOrder', '')).strip() or None
                word_order = str(item.get('wordOrder', '')).strip() or None

                tag_vo = TagModel(
                    deviceId=device_id, tagName=tag_name,
                    registerType=register_type.upper(), registerAddress=register_address,
                    dataType=data_type.upper(), unit=unit, description=description,
                    status='0', sortOrder=sort_order,
                    transformType=transform_type,
                    transformSlopeA=transform_slope_a,
                    transformOffsetB=transform_offset_b,
                    engUnit=eng_unit,
                    reportDeadbandMs=report_deadband_ms,
                    reportForceIntervalMs=report_force_interval_ms,
                    bitOffset=bit_offset,
                    byteOrder=byte_order,
                    wordOrder=word_order,
                    createBy=create_by,
                    createTime=now,
                    updateBy=create_by,
                    updateTime=now,
                )
                valid_tags.append(tag_vo)
            except Exception as e:
                errors.append(TagImportError(row=row_num, reason=str(e)))

        try:
            if valid_tags:
                success_count = await TagDao.batch_add_tags(query_db, valid_tags)
                await cls._log_change(
                    query_db, 'add', 'tag', 0, f'批量导入点位(device={device_info.device_name})', create_by,
                    after_value={'successCount': success_count, 'failCount': len(errors)},
                )
                await query_db.commit()
                logger.info(f'点位导入完成：成功 {success_count}，失败 {len(errors)}')
            elif errors:
                await query_db.rollback()
            else:
                await query_db.rollback()
        except Exception:
            await query_db.rollback()
            raise

        return TagImportResult(success_count=success_count, fail_count=len(errors), errors=errors)

    @staticmethod
    def get_tag_template_services() -> bytes:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = '点位导入模板'
        headers = [
            '点位名称', '寄存器类型', '寄存器地址', '数据类型', '单位', '描述', '排序号',
            '换算类型', '斜率/乘数a', '偏移量b', '工程单位', '死区(数值)', '强制间隔(ms)',
            '位偏移', '字节序', '字序',
        ]
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = 16

        dv_reg = DataValidation(type='list', formula1='"D,W,X,Y,M"', allow_blank=True)
        dv_reg.error = '请选择有效的寄存器类型'
        ws.add_data_validation(dv_reg)
        dv_reg.add('B2:B1048576')

        dv_dt = DataValidation(type='list', formula1='"INT16,INT32,FLOAT,BIT,UINT16,UINT32,BOOL,DOUBLE"', allow_blank=True)
        dv_dt.error = '请选择有效的数据类型'
        ws.add_data_validation(dv_dt)
        dv_dt.add('D2:D1048576')

        dv_bo = DataValidation(type='list', formula1='"LITTLE_ENDIAN,BIG_ENDIAN"', allow_blank=True)
        dv_bo.error = '请选择有效的字节序'
        ws.add_data_validation(dv_bo)
        dv_bo.add('N2:N1048576')

        dv_wo = DataValidation(type='list', formula1='"LOW_FIRST,HIGH_FIRST"', allow_blank=True)
        dv_wo.error = '请选择有效的字序'
        ws.add_data_validation(dv_wo)
        dv_wo.add('O2:O1048576')

        dv_trans = DataValidation(type='list', formula1='"none,linear,slope_offset"', allow_blank=True)
        dv_trans.error = '请选择有效的换算类型'
        ws.add_data_validation(dv_trans)
        dv_trans.add('H2:H1048576')

        binary_data = io.BytesIO()
        wb.save(binary_data)
        binary_data.seek(0)
        return binary_data.getvalue()


    @classmethod
    async def set_tag_status(
        cls, query_db: AsyncSession, tag_id: int, status: str,
        update_by: str = '', update_time=None,
    ) -> None:
        tag_info = await TagDao.get_tag_detail_by_id(query_db, tag_id)
        if not tag_info:
            raise ServiceException(message=f'PLC点位不存在：ID={tag_id}')
        if status not in {'0', '1'}:
            raise ServiceException(message=f'无效的点位状态：{status}，仅支持 0（启用）或 1（停用）')
        if status == '1':
            await cls._assert_not_calc_source(query_db, [tag_id], action='停用')
        update_dict: dict = {'id': tag_id, 'status': status}
        if update_by:
            update_dict['update_by'] = update_by
        if update_time:
            update_dict['update_time'] = update_time
        await TagDao.edit_tag_dao(query_db, update_dict)
        await cls._log_change(
            query_db, 'enable' if status == '0' else 'disable', 'tag', tag_id,
            tag_info.tag_name, change_by=update_by,
            before_value={'status': tag_info.status},
            after_value={'status': status},
        )
        await query_db.commit()

    # ==================== 换算工具 ====================

    @staticmethod
    def compute_eng_value(raw_value: float, tag: dict | TagModel) -> tuple[float | None, str]:
        """
        根据点位换算配置，将原始值转为工程值

        :param raw_value: 从PLC读取到的原始数值
        :param tag: 点位配置（dict 或 TagModel）
        :return: (工程值, 数据质量码) — 质量码: GOOD / BAD / UNCERTAIN
        """
        if raw_value is None:
            return None, 'BAD'

        # 获取配置
        if isinstance(tag, dict):
            transform_type = tag.get('transformType', tag.get('transform_type', 'none'))
            slope_a = tag.get('transformSlopeA', tag.get('transform_slope_a', 1.0)) or 1.0
            offset_b = tag.get('transformOffsetB', tag.get('transform_offset_b', 0.0)) or 0.0
            raw_min = tag.get('rawValueMin', tag.get('raw_value_min'))
            raw_max = tag.get('rawValueMax', tag.get('raw_value_max'))
            eng_min = tag.get('engValueMin', tag.get('eng_value_min'))
            eng_max = tag.get('engValueMax', tag.get('eng_value_max'))
        else:
            transform_type = tag.transform_type or 'none'
            slope_a = tag.transform_slope_a or 1.0
            offset_b = tag.transform_offset_b or 0.0
            raw_min = tag.raw_value_min
            raw_max = tag.raw_value_max
            eng_min = tag.eng_value_min
            eng_max = tag.eng_value_max

        # 检查原始值是否在有效范围内
        if raw_min is not None and raw_max is not None:
            if raw_value < raw_min or raw_value > raw_max:
                return None, 'UNCERTAIN'

        # 按类型换算
        transform_type = (transform_type or 'none').lower()
        if transform_type == 'none':
            return raw_value, 'GOOD'
        elif transform_type == 'linear':
            # y = a × x + b
            return slope_a * raw_value + offset_b, 'GOOD'
        elif transform_type == 'slope_offset':
            # y = (x - raw_min) / (raw_max - raw_min) × (eng_max - eng_min) + eng_min
            if None in (raw_min, raw_max, eng_min, eng_max):
                return raw_value, 'UNCERTAIN'  # 缺少量程参数
            if raw_max == raw_min:
                return eng_min, 'UNCERTAIN'  # 量程为零宽度，说明配置有误
            ratio = (raw_value - raw_min) / (raw_max - raw_min)
            return eng_min + ratio * (eng_max - eng_min), 'GOOD'
        else:
            return raw_value, 'UNCERTAIN'  # 未知换算类型

    @staticmethod
    def compute_from_tag(raw_value: float, transform_type: str, slope_a: float = 1.0,
                         offset_b: float = 0.0, **kwargs) -> tuple[float | None, str]:
        """
        轻量版：直接用参数计算（供 Node-RED 侧参考逻辑）

        :param raw_value: 原始值
        :param transform_type: none / linear / slope_offset
        :param slope_a: 斜率
        :param offset_b: 偏移
        :return: (工程值, 质量码)
        """
        if raw_value is None:
            return None, 'BAD'
        t = (transform_type or 'none').lower()
        if t == 'none':
            return raw_value, 'GOOD'
        elif t == 'linear':
            return slope_a * raw_value + offset_b, 'GOOD'
        elif t == 'slope_offset':
            # 量程映射：raw_min→eng_min, raw_max→eng_max
            raw_min = kwargs.get('raw_value_min')
            raw_max = kwargs.get('raw_value_max')
            eng_min = kwargs.get('eng_value_min')
            eng_max = kwargs.get('eng_value_max')
            if None in (raw_min, raw_max, eng_min, eng_max) or raw_min == raw_max:
                return raw_value, 'UNCERTAIN'
            ratio = (raw_value - raw_min) / (raw_max - raw_min)
            return eng_min + ratio * (eng_max - eng_min), 'GOOD'
        else:
            return raw_value, 'UNCERTAIN'

    # ==================== 导出 ====================

    @classmethod
    async def export_tag_list_services(cls, query_db: AsyncSession, query_object: TagPageQueryModel) -> bytes:
        """
        导出点位列表为 Excel（含设备名称列）

        :param query_db: orm对象
        :param query_object: 查询参数
        :return: Excel 二进制数据
        """
        import io
        from openpyxl import Workbook

        # 使用 write_only 模式 + 分页写入，避免大数据量 OOM
        wb = Workbook(write_only=True)
        ws = wb.create_sheet('PLC点位清单')

        tag_headers = [
            '设备名称', '点位名称', '寄存器类型', '寄存器地址',
            '数据类型', '单位', '描述', '状态', '排序号',
            '位偏移', '字节序', '字序',
        ]
        tag_keys = [
            'deviceName', 'tagName', 'registerType', 'registerAddress',
            'dataType', 'unit', 'description', 'status', 'sortOrder',
            'bitOffset', 'byteOrder', 'wordOrder',
        ]

        # 分页流式写入：每批500条
        page_size = 500
        page_num = 1
        first_page = True

        while True:
            query_object.page_num = page_num
            query_object.page_size = page_size
            tag_result = await TagDao.get_tag_global_list(query_db, query_object, is_page=True)
            page_data = tag_result.rows if hasattr(tag_result, 'rows') else []

            if not page_data:
                break

            for tag in page_data:
                if first_page:
                    ws.append(tag_headers)
                    first_page = False
                row_data = []
                for key in tag_keys:
                    val = tag.get(key, '') if isinstance(tag, dict) else ''
                    if key == 'status':
                        val = '启用' if val == '0' else '停用' if val == '1' else val
                    row_data.append(val)
                ws.append(row_data)

            if len(page_data) < page_size:
                break
            page_num += 1

        binary_data = io.BytesIO()
        wb.save(binary_data)
        binary_data.seek(0)
        return binary_data.getvalue()
